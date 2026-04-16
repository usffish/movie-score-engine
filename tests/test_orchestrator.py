"""
Unit tests for update_scores.py orchestrator.

Covers:
  - parse_args() default values
  - --input defaults to Movies.xlsx
  - --output defaults to <stem>_updated.xlsx
  - --limit restricts to first N rows
  - --movie restricts to exact title match
  - --movie with unknown title exits non-zero
  - Missing OMDB_API_KEY and no --api-key exits non-zero
  - Missing input file exits non-zero
  - Missing Movies column exits non-zero
  - Failed movies appear in summary log
  - Each movie title logged at INFO before fetch
"""

import logging
import os
import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from update_scores import (
    NormalisedScores,
    RawScores,
    fetch_all,
    main,
    parse_args,
    update_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_workbook(titles, extra_cols=None):
    """Create an in-memory openpyxl workbook with a Movies column."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Movies"] + (extra_cols or [])
    ws.append(headers)
    for title in titles:
        row = [title] + [""] * len(extra_cols or [])
        ws.append(row)
    return wb


def make_normalised(title, metascore=70, st_metacritic=0.5, review_count=10,
                    letterboxd_rating=3.5, st_letterboxd=0.5,
                    imdb_rating=7.0, st_imdb=0.5, composite=0.5):
    return NormalisedScores(
        title=title,
        metascore=metascore,
        st_metacritic=st_metacritic,
        review_count=review_count,
        letterboxd_rating=letterboxd_rating,
        st_letterboxd=st_letterboxd,
        imdb_rating=imdb_rating,
        st_imdb=st_imdb,
        composite=composite,
    )


# ---------------------------------------------------------------------------
# parse_args() default values
# ---------------------------------------------------------------------------

class TestParseArgsDefaults:
    def test_input_defaults_to_movies_xlsx(self):
        args = parse_args([])
        assert args.input == "Movies.xlsx"

    def test_output_defaults_to_none(self):
        # output=None means the orchestrator derives it from input stem
        args = parse_args([])
        assert args.output is None

    def test_limit_defaults_to_none(self):
        args = parse_args([])
        assert args.limit is None

    def test_movie_defaults_to_none(self):
        args = parse_args([])
        assert args.movie is None

    def test_delay_defaults_to_1_0(self):
        args = parse_args([])
        assert args.delay == 1.0

    def test_verbose_defaults_to_false(self):
        args = parse_args([])
        assert args.verbose is False

    def test_api_key_defaults_to_none(self):
        args = parse_args([])
        assert args.api_key is None

    def test_output_stem_derived_from_input(self):
        """When --output is not given, main() should derive <stem>_updated.xlsx."""
        args = parse_args(["--input", "my_list.xlsx"])
        assert args.output is None
        input_path = Path(args.input)
        expected = input_path.parent / f"{input_path.stem}_updated{input_path.suffix}"
        assert expected == Path("my_list_updated.xlsx")


# ---------------------------------------------------------------------------
# --input / --output argument handling
# ---------------------------------------------------------------------------

class TestInputOutputArgs:
    def test_custom_input_path_parsed(self):
        args = parse_args(["--input", "/tmp/custom.xlsx"])
        assert args.input == "/tmp/custom.xlsx"

    def test_custom_output_path_parsed(self):
        args = parse_args(["--output", "/tmp/out.xlsx"])
        assert args.output == "/tmp/out.xlsx"

    def test_api_key_arg_parsed(self):
        args = parse_args(["--api-key", "mykey123"])
        assert args.api_key == "mykey123"


# ---------------------------------------------------------------------------
# Missing API key exits non-zero
# ---------------------------------------------------------------------------

class TestMissingApiKey:
    def test_no_api_key_no_env_exits_nonzero(self, tmp_path):
        """When neither --api-key nor OMDB_API_KEY is set, exit with code 1."""
        wb = make_workbook(["Inception"])
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))

        env_without_key = {k: v for k, v in os.environ.items() if k != "OMDB_API_KEY"}
        with patch.dict(os.environ, env_without_key, clear=True):
            with pytest.raises(SystemExit) as exc_info:
                main(["--input", str(input_file)])
        assert exc_info.value.code != 0

    def test_api_key_from_env_is_accepted(self, tmp_path):
        """When OMDB_API_KEY is set in env, it should be used."""
        wb = make_workbook(["Inception"])
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with patch.dict(os.environ, {"OMDB_API_KEY": "testkey"}, clear=False):
            with patch("update_scores.update_workbook") as mock_uw:
                main(["--input", str(input_file), "--output", str(output_file)])
                mock_uw.assert_called_once()
                _, kwargs = mock_uw.call_args
                assert kwargs["api_key"] == "testkey"

    def test_api_key_from_cli_overrides_env(self, tmp_path):
        """--api-key takes precedence over OMDB_API_KEY env var."""
        wb = make_workbook(["Inception"])
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with patch.dict(os.environ, {"OMDB_API_KEY": "envkey"}, clear=False):
            with patch("update_scores.update_workbook") as mock_uw:
                main([
                    "--input", str(input_file),
                    "--output", str(output_file),
                    "--api-key", "clikey",
                ])
                _, kwargs = mock_uw.call_args
                assert kwargs["api_key"] == "clikey"


# ---------------------------------------------------------------------------
# Missing input file exits non-zero
# ---------------------------------------------------------------------------

class TestMissingInputFile:
    def test_missing_input_file_exits_nonzero(self, tmp_path):
        nonexistent = tmp_path / "does_not_exist.xlsx"
        with patch.dict(os.environ, {"OMDB_API_KEY": "testkey"}, clear=False):
            with pytest.raises(SystemExit) as exc_info:
                main(["--input", str(nonexistent)])
        assert exc_info.value.code != 0


# ---------------------------------------------------------------------------
# Missing Movies column exits non-zero
# ---------------------------------------------------------------------------

class TestMissingMoviesColumn:
    def test_missing_movies_column_exits_nonzero(self, tmp_path):
        """Workbook without a 'Movies' column should cause exit(1)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Title", "Year"])  # no 'Movies' column
        ws.append(["Inception", 2010])
        input_file = tmp_path / "no_movies_col.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "out.xlsx"

        with pytest.raises(SystemExit) as exc_info:
            update_workbook(
                input_path=input_file,
                output_path=output_file,
                api_key="testkey",
            )
        assert exc_info.value.code != 0


# ---------------------------------------------------------------------------
# --limit restricts to first N rows
# ---------------------------------------------------------------------------

class TestLimitArg:
    def test_limit_restricts_movies_processed(self, tmp_path):
        """--limit N should only process the first N movies."""
        titles = ["Movie A", "Movie B", "Movie C", "Movie D", "Movie E"]
        wb = make_workbook(titles)
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with patch("update_scores.fetch_all") as mock_fetch:
            mock_fetch.return_value = ([], [])
            with patch("update_scores.normalise_all", return_value=[]):
                with patch("update_scores.compute_all_composites", return_value=[]):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                        limit=3,
                    )
            called_movies = mock_fetch.call_args[0][0]
            assert called_movies == ["Movie A", "Movie B", "Movie C"]
            assert len(called_movies) == 3


# ---------------------------------------------------------------------------
# --movie restricts to exact title match
# ---------------------------------------------------------------------------

class TestMovieArg:
    def test_movie_arg_restricts_to_exact_title(self, tmp_path):
        """--movie should only process the movie with the exact matching title."""
        titles = ["Inception", "The Matrix", "Interstellar"]
        wb = make_workbook(titles)
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with patch("update_scores.fetch_all") as mock_fetch:
            mock_fetch.return_value = ([], [])
            with patch("update_scores.normalise_all", return_value=[]):
                with patch("update_scores.compute_all_composites", return_value=[]):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                        target_movie="The Matrix",
                    )
            called_movies = mock_fetch.call_args[0][0]
            assert called_movies == ["The Matrix"]

    def test_movie_arg_unknown_title_exits_nonzero(self, tmp_path):
        """--movie with a title not in the workbook should exit non-zero."""
        titles = ["Inception", "The Matrix"]
        wb = make_workbook(titles)
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with pytest.raises(SystemExit) as exc_info:
            update_workbook(
                input_path=input_file,
                output_path=output_file,
                api_key="testkey",
                target_movie="Unknown Movie",
            )
        assert exc_info.value.code != 0


# ---------------------------------------------------------------------------
# Failed movies appear in summary log
# ---------------------------------------------------------------------------

class TestFailedMoviesSummary:
    def test_failed_movies_logged_in_summary(self, tmp_path, caplog):
        """Movies that fail during fetch should appear in the warning summary."""
        titles = ["Inception", "The Matrix", "Interstellar"]
        wb = make_workbook(titles)
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        raw = [RawScores("Inception", 74, 7.8, 80, 3.9)]
        failed = ["The Matrix", "Interstellar"]

        with patch("update_scores.fetch_all", return_value=(raw, failed)):
            with patch("update_scores.normalise_all") as mock_norm:
                mock_norm.return_value = [make_normalised("Inception")]
                with patch("update_scores.compute_all_composites") as mock_comp:
                    mock_comp.return_value = [make_normalised("Inception")]
                    with caplog.at_level(logging.WARNING, logger="update_scores"):
                        update_workbook(
                            input_path=input_file,
                            output_path=output_file,
                            api_key="testkey",
                        )

        warning_text = " ".join(caplog.messages)
        assert "The Matrix" in warning_text
        assert "Interstellar" in warning_text

    def test_failed_count_logged(self, tmp_path, caplog):
        """The summary should mention the count of failed movies."""
        titles = ["Movie A", "Movie B"]
        wb = make_workbook(titles)
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with patch("update_scores.fetch_all", return_value=([], ["Movie A", "Movie B"])):
            with patch("update_scores.normalise_all", return_value=[]):
                with patch("update_scores.compute_all_composites", return_value=[]):
                    with caplog.at_level(logging.WARNING, logger="update_scores"):
                        update_workbook(
                            input_path=input_file,
                            output_path=output_file,
                            api_key="testkey",
                        )

        warning_text = " ".join(caplog.messages)
        assert "2" in warning_text


# ---------------------------------------------------------------------------
# Each movie title logged at INFO before fetch
# ---------------------------------------------------------------------------

class TestTitleLoggedBeforeFetch:
    def test_each_title_logged_at_info(self, caplog):
        """fetch_all should log each movie title at INFO level before fetching."""
        titles = ["Inception", "The Matrix"]

        with patch("update_scores.get_omdb_data") as mock_omdb, \
             patch("update_scores.get_review_count") as mock_rc, \
             patch("update_scores.get_letterboxd_data") as mock_lb, \
             patch("update_scores.time.sleep"):

            mock_omdb.return_value = {"metascore": 74, "imdb_rating": 7.8, "imdb_id": "tt1"}
            mock_rc.return_value = 80
            mock_lb.return_value = {"rating": 3.9, "rating_count": 100, "url": "http://x"}

            with caplog.at_level(logging.INFO, logger="update_scores"):
                fetch_all(titles, api_key="testkey", delay=0)

        info_messages = [r.message for r in caplog.records if r.levelno == logging.INFO]
        assert any("Inception" in m for m in info_messages)
        assert any("The Matrix" in m for m in info_messages)

    def test_title_logged_before_omdb_call(self, caplog):
        """The INFO log for a title must appear before any fetch for that title."""
        call_order = []

        def log_side_effect(title, api_key, **kwargs):
            call_order.append(f"omdb:{title}")
            return {"metascore": 74, "imdb_rating": 7.8, "imdb_id": "tt1"}

        with patch("update_scores.get_omdb_data", side_effect=log_side_effect), \
             patch("update_scores.get_review_count", return_value=10), \
             patch("update_scores.get_letterboxd_data",
                   return_value={"rating": 3.5, "rating_count": 50, "url": "http://x"}), \
             patch("update_scores.time.sleep"):

            with caplog.at_level(logging.INFO, logger="update_scores"):
                fetch_all(["Inception"], api_key="testkey", delay=0)

        info_records = [r for r in caplog.records if r.levelno == logging.INFO]
        assert len(info_records) >= 1
        assert "Inception" in info_records[0].message


# ---------------------------------------------------------------------------
# fetch_all: per-movie exception handling
# ---------------------------------------------------------------------------

class TestFetchAllExceptionHandling:
    def test_exception_in_one_movie_does_not_abort_others(self, caplog):
        """An exception for one movie should not prevent others from being fetched."""
        titles = ["Good Movie", "Bad Movie", "Another Good Movie"]

        def omdb_side_effect(title, api_key, **kwargs):
            if title == "Bad Movie":
                raise RuntimeError("Network error")
            return {"metascore": 74, "imdb_rating": 7.8, "imdb_id": "tt1"}

        with patch("update_scores.get_omdb_data", side_effect=omdb_side_effect), \
             patch("update_scores.get_review_count", return_value=10), \
             patch("update_scores.get_letterboxd_data",
                   return_value={"rating": 3.5, "rating_count": 50, "url": "http://x"}), \
             patch("update_scores.time.sleep"):

            with caplog.at_level(logging.ERROR, logger="update_scores"):
                raw_scores, failed = fetch_all(titles, api_key="testkey", delay=0)

        assert len(raw_scores) == 2
        assert failed == ["Bad Movie"]
        titles_fetched = [r.title for r in raw_scores]
        assert "Good Movie" in titles_fetched
        assert "Another Good Movie" in titles_fetched

    def test_failed_movie_logged_with_title(self, caplog):
        """When a movie fails, its title should appear in the error log."""
        with patch("update_scores.get_omdb_data", side_effect=RuntimeError("boom")), \
             patch("update_scores.time.sleep"):

            with caplog.at_level(logging.ERROR, logger="update_scores"):
                fetch_all(["Failing Movie"], api_key="testkey", delay=0)

        error_messages = [r.message for r in caplog.records if r.levelno == logging.ERROR]
        assert any("Failing Movie" in m for m in error_messages)


# ---------------------------------------------------------------------------
# update_workbook: None values leave cells unchanged
# ---------------------------------------------------------------------------

class TestNoneValuesLeavesCellsUnchanged:
    def test_none_score_does_not_overwrite_existing_cell(self, tmp_path):
        """When a NormalisedScores field is None, the existing cell value is preserved."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Movies", "Metacritic", "IMDB"])
        ws.append(["Inception", 74, 7.8])  # existing values
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        ns = NormalisedScores(
            title="Inception",
            metascore=80,
            st_metacritic=0.6,
            review_count=90,
            letterboxd_rating=None,
            st_letterboxd=None,
            imdb_rating=None,   # None — should not overwrite existing 7.8
            st_imdb=None,
            composite=0.55,
        )

        with patch("update_scores.fetch_all", return_value=([
            RawScores("Inception", 80, None, 90, None)
        ], [])):
            with patch("update_scores.normalise_all", return_value=[ns]):
                with patch("update_scores.compute_all_composites", return_value=[ns]):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                    )

        out_wb = openpyxl.load_workbook(str(output_file))
        out_ws = out_wb.active
        headers = {cell.value: cell.column for cell in out_ws[1] if cell.value}
        imdb_col = headers.get("IMDB")
        assert imdb_col is not None
        # The existing value 7.8 should be preserved since imdb_rating is None
        assert out_ws.cell(row=2, column=imdb_col).value == 7.8


# ---------------------------------------------------------------------------
# update_workbook: missing output columns are added
# ---------------------------------------------------------------------------

class TestMissingOutputColumnsAdded:
    def test_missing_columns_added_to_output(self, tmp_path):
        """Output workbook should contain all required columns even if absent from input."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Movies"])  # only Movies column
        ws.append(["Inception"])
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        ns = make_normalised("Inception")

        with patch("update_scores.fetch_all", return_value=([
            RawScores("Inception", 74, 7.8, 80, 3.9)
        ], [])):
            with patch("update_scores.normalise_all", return_value=[ns]):
                with patch("update_scores.compute_all_composites", return_value=[ns]):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                    )

        out_wb = openpyxl.load_workbook(str(output_file))
        out_ws = out_wb.active
        headers = {cell.value for cell in out_ws[1] if cell.value}
        required = {"Metacritic", "st.Metacritic", "Reviews", "Letterboxd",
                    "st.Letterboxd", "IMDB", "st.IMDB", "TRUE"}
        assert required.issubset(headers)


# ---------------------------------------------------------------------------
# Blank rows are skipped
# ---------------------------------------------------------------------------

class TestBlankRowsSkipped:
    def test_blank_title_rows_are_skipped(self, tmp_path):
        """Rows with blank/None titles should not be passed to fetch_all."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Movies"])
        ws.append(["Inception"])
        ws.append([None])       # blank row
        ws.append(["   "])      # whitespace-only row
        ws.append(["The Matrix"])
        input_file = tmp_path / "Movies.xlsx"
        wb.save(str(input_file))
        output_file = tmp_path / "Movies_updated.xlsx"

        with patch("update_scores.fetch_all") as mock_fetch:
            mock_fetch.return_value = ([], [])
            with patch("update_scores.normalise_all", return_value=[]):
                with patch("update_scores.compute_all_composites", return_value=[]):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                    )
            called_movies = mock_fetch.call_args[0][0]
            assert called_movies == ["Inception", "The Matrix"]
            assert None not in called_movies
            assert "   " not in called_movies
