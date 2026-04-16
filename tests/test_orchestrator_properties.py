"""
Property-based tests for the update_scores orchestrator.

Design references:
  Property 6  — Input workbook is never modified
    **Validates: Requirements 7.2**
  Property 7  — All required output columns are present
    **Validates: Requirements 7.3, 7.4**
  Property 8  — None values leave cells unchanged
    **Validates: Requirements 7.5**
  Property 9  — Error resilience — processing continues after failure
    **Validates: Requirements 9.1**
  Property 13 — Blank and whitespace-only movie rows are skipped
    **Validates: Requirements 1.3**
"""

import tempfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
from hypothesis import given, settings
from hypothesis import strategies as st

from update_scores import (
    NormalisedScores,
    RawScores,
    fetch_all,
    update_workbook,
)

# ---------------------------------------------------------------------------
# Shared strategies
# ---------------------------------------------------------------------------

# Valid movie title: printable ASCII characters only (safe for openpyxl cells)
movie_title_strategy = st.text(
    alphabet=st.characters(
        whitelist_categories=("Lu", "Ll", "Nd", "Zs"),
        whitelist_characters="!@#$%^&*()-_=+[]{}|;:',.<>?/`~",
    ),
    min_size=1,
    max_size=50,
).filter(lambda s: s.strip() != "" and s == s.strip())

# List of valid movie titles
movie_list_strategy = st.lists(movie_title_strategy, min_size=1, max_size=10)

# Existing cell values: use integers to avoid float precision round-trip issues
# openpyxl stores floats but may lose precision; integers are exact
existing_score_strategy = st.integers(min_value=0, max_value=100)

# Blank / whitespace-only / None values for movie rows
blank_value_strategy = st.one_of(
    st.none(),
    st.just(""),
    st.text(alphabet=" \t\n", min_size=1, max_size=5),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = {
    "Metacritic", "st.Metacritic", "Reviews",
    "Letterboxd", "st.Letterboxd", "IMDB", "st.IMDB", "TRUE",
}


def _save_workbook(path: Path, titles: list) -> None:
    """Save a minimal workbook with a Movies column to *path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Movies"])
    for title in titles:
        ws.append([title])
    wb.save(str(path))


def _make_normalised(title: str) -> NormalisedScores:
    """Return a fully-populated NormalisedScores for *title*."""
    return NormalisedScores(
        title=title,
        metascore=70,
        st_metacritic=0.5,
        review_count=10,
        letterboxd_rating=3.5,
        st_letterboxd=0.5,
        imdb_rating=7.0,
        st_imdb=0.5,
        composite=0.5,
    )


# ---------------------------------------------------------------------------
# Property 6: Input workbook is never modified
# **Validates: Requirements 7.2**
# ---------------------------------------------------------------------------

@given(titles=movie_list_strategy)
@settings(max_examples=50)
def test_property6_input_workbook_unchanged(titles):
    """
    Property 6: For any input workbook, after running update_workbook the
    input file's contents are byte-for-byte identical to its state before
    the run.

    **Validates: Requirements 7.2**
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_file = tmp_path / "input.xlsx"
        output_file = tmp_path / "output.xlsx"

        _save_workbook(input_file, titles)

        # Record the input file's bytes before the run
        before_bytes = input_file.read_bytes()

        normalised = [_make_normalised(t) for t in titles]

        with patch("update_scores.fetch_all", return_value=([
            RawScores(t, 70, 7.0, 10, 3.5) for t in titles
        ], [])):
            with patch("update_scores.normalise_all", return_value=normalised):
                with patch("update_scores.compute_all_composites", return_value=normalised):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                        delay=0.0,
                    )

        after_bytes = input_file.read_bytes()

        assert before_bytes == after_bytes, (
            "Input workbook was modified during update_workbook run "
            f"(titles={titles})"
        )


# ---------------------------------------------------------------------------
# Property 7: All required output columns are present
# **Validates: Requirements 7.3, 7.4**
# ---------------------------------------------------------------------------

@given(titles=movie_list_strategy)
@settings(max_examples=50)
def test_property7_all_required_columns_present(titles):
    """
    Property 7: For any movie list processed by the orchestrator, the output
    workbook contains all 8 required columns: Metacritic, st.Metacritic,
    Reviews, Letterboxd, st.Letterboxd, IMDB, st.IMDB, TRUE.

    **Validates: Requirements 7.3, 7.4**
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_file = tmp_path / "input.xlsx"
        output_file = tmp_path / "output.xlsx"

        _save_workbook(input_file, titles)

        normalised = [_make_normalised(t) for t in titles]

        with patch("update_scores.fetch_all", return_value=([
            RawScores(t, 70, 7.0, 10, 3.5) for t in titles
        ], [])):
            with patch("update_scores.normalise_all", return_value=normalised):
                with patch("update_scores.compute_all_composites", return_value=normalised):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                        delay=0.0,
                    )

        out_wb = openpyxl.load_workbook(str(output_file))
        out_ws = out_wb.active
        headers = {cell.value for cell in out_ws[1] if cell.value is not None}

        missing = REQUIRED_COLUMNS - headers
        assert not missing, (
            f"Output workbook is missing required columns: {missing} "
            f"(titles={titles})"
        )


# ---------------------------------------------------------------------------
# Property 8: None values leave cells unchanged
# **Validates: Requirements 7.5**
# ---------------------------------------------------------------------------

@given(
    title=movie_title_strategy,
    existing_imdb=existing_score_strategy,
    existing_letterboxd=existing_score_strategy,
)
@settings(max_examples=50)
def test_property8_none_scores_do_not_overwrite_existing_cells(
    title, existing_imdb, existing_letterboxd
):
    """
    Property 8: For any input workbook with existing cell values, when a
    scraper returns None for a score, the corresponding cell in the output
    workbook retains its original value.

    Uses integer scores to avoid float precision round-trip issues with openpyxl.

    **Validates: Requirements 7.5**
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_file = tmp_path / "input.xlsx"
        output_file = tmp_path / "output.xlsx"

        # Build a workbook that already has IMDB and Letterboxd values
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Movies", "IMDB", "Letterboxd"])
        ws.append([title, existing_imdb, existing_letterboxd])
        wb.save(str(input_file))

        # NormalisedScores with None for imdb_rating and letterboxd_rating
        # so those cells must NOT be overwritten
        ns = NormalisedScores(
            title=title,
            metascore=70,
            st_metacritic=0.5,
            review_count=10,
            letterboxd_rating=None,   # None — must not overwrite existing_letterboxd
            st_letterboxd=None,
            imdb_rating=None,         # None — must not overwrite existing_imdb
            st_imdb=None,
            composite=0.5,
        )

        with patch("update_scores.fetch_all", return_value=([
            RawScores(title, 70, None, 10, None)
        ], [])):
            with patch("update_scores.normalise_all", return_value=[ns]):
                with patch("update_scores.compute_all_composites", return_value=[ns]):
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                        delay=0.0,
                    )

        out_wb = openpyxl.load_workbook(str(output_file))
        out_ws = out_wb.active
        headers = {cell.value: cell.column for cell in out_ws[1] if cell.value is not None}

        imdb_col = headers.get("IMDB")
        lb_col = headers.get("Letterboxd")

        assert imdb_col is not None, "IMDB column not found in output"
        assert lb_col is not None, "Letterboxd column not found in output"

        actual_imdb = out_ws.cell(row=2, column=imdb_col).value
        actual_lb = out_ws.cell(row=2, column=lb_col).value

        assert actual_imdb == existing_imdb, (
            f"IMDB cell was overwritten: expected {existing_imdb}, got {actual_imdb} "
            f"(title={title!r})"
        )
        assert actual_lb == existing_letterboxd, (
            f"Letterboxd cell was overwritten: expected {existing_letterboxd}, "
            f"got {actual_lb} (title={title!r})"
        )


# ---------------------------------------------------------------------------
# Property 9: Error resilience — processing continues after failure
# **Validates: Requirements 9.1**
# ---------------------------------------------------------------------------

@given(titles=movie_list_strategy)
@settings(max_examples=50)
def test_property9_remaining_movies_processed_after_failure(titles):
    """
    Property 9: For any movie list where a subset of movies raise exceptions
    during fetching, the orchestrator processes all remaining movies and the
    set of successfully processed movies equals the full list minus the
    failing movies.

    **Validates: Requirements 9.1**
    """
    # Mark every other title as "failing" to ensure a mix
    failing = set(titles[::2])
    succeeding = [t for t in titles if t not in failing]

    def omdb_side_effect(title, api_key, **kwargs):
        if title in failing:
            raise RuntimeError(f"Simulated network error for '{title}'")
        return {"metascore": 70, "imdb_rating": 7.0, "imdb_id": "tt0000001"}

    with patch("update_scores.get_omdb_data", side_effect=omdb_side_effect), \
         patch("update_scores.get_review_count", return_value=10), \
         patch("update_scores.get_letterboxd_data",
               return_value={"rating": 3.5, "rating_count": 100, "url": "http://x"}), \
         patch("update_scores.time.sleep"):

        raw_scores, failed = fetch_all(titles, api_key="testkey", delay=0.0)

    succeeded_titles = {r.title for r in raw_scores}
    failed_titles = set(failed)

    # Every non-failing title must appear in raw_scores
    for t in succeeding:
        assert t in succeeded_titles, (
            f"Title '{t}' should have succeeded but is missing from raw_scores "
            f"(titles={titles}, failing={failing})"
        )

    # Every failing title must appear in failed list
    for t in failing:
        assert t in failed_titles, (
            f"Title '{t}' should be in failed list but is missing "
            f"(titles={titles})"
        )

    # The union of succeeded and failed must equal the full input set
    assert succeeded_titles | failed_titles == set(titles), (
        f"succeeded union failed != full title set "
        f"(succeeded={succeeded_titles}, failed={failed_titles}, titles={titles})"
    )


# ---------------------------------------------------------------------------
# Property 13: Blank and whitespace-only movie rows are skipped
# **Validates: Requirements 1.3**
# ---------------------------------------------------------------------------

@given(
    valid_titles=movie_list_strategy,
    blank_values=st.lists(blank_value_strategy, min_size=1, max_size=5),
)
@settings(max_examples=50)
def test_property13_blank_rows_excluded_from_processing(valid_titles, blank_values):
    """
    Property 13: For any workbook containing rows where the Movies column
    value is None, empty string, or whitespace-only, those rows are skipped
    without raising and do not appear in the list passed to fetch_all.

    **Validates: Requirements 1.3**
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_file = tmp_path / "input.xlsx"
        output_file = tmp_path / "output.xlsx"

        # Interleave valid titles and blank values in the workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Movies"])
        for i, title in enumerate(valid_titles):
            ws.append([title])
            # Insert a blank row after each valid title (if blanks remain)
            if i < len(blank_values):
                ws.append([blank_values[i]])
        # Append any remaining blank values
        for bv in blank_values[len(valid_titles):]:
            ws.append([bv])
        wb.save(str(input_file))

        captured_movies = []

        def capture_fetch_all(movies, api_key, delay=0.0, verbose=False):
            captured_movies.extend(movies)
            return [], []

        with patch("update_scores.fetch_all", side_effect=capture_fetch_all):
            with patch("update_scores.normalise_all", return_value=[]):
                with patch("update_scores.compute_all_composites", return_value=[]):
                    # Must not raise
                    update_workbook(
                        input_path=input_file,
                        output_path=output_file,
                        api_key="testkey",
                        delay=0.0,
                    )

        # None, empty strings, and whitespace-only strings must not appear
        for movie in captured_movies:
            assert movie is not None, (
                f"None value was passed to fetch_all (valid_titles={valid_titles}, "
                f"blank_values={blank_values})"
            )
            assert movie.strip() != "", (
                f"Blank/whitespace title {movie!r} was passed to fetch_all "
                f"(valid_titles={valid_titles}, blank_values={blank_values})"
            )

        # All valid titles must be present
        for title in valid_titles:
            assert title in captured_movies, (
                f"Valid title {title!r} was not passed to fetch_all "
                f"(captured={captured_movies})"
            )
