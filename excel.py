"""
excel.py
========
Workbook I/O, header management, and stability-tracking helpers.
"""

import logging
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

from scoring import NormalisedScores, RawScores

logger = logging.getLogger(__name__)

EXPECTED_HEADERS = [
    "Movies", "Metacritic", "st.Metacritic", "Reviews",
    "Letterboxd", "st.Letterboxd", "IMDB", "st.IMDB", "TRUE",
    "LastUpdated", "StableWeeks",
]

_TABLE_NAME = "Table1"
_TABLE_CORE_COLS = [
    "Movies", "Metacritic", "st.Metacritic", "Reviews",
    "Letterboxd", "st.Letterboxd", "IMDB", "st.IMDB", "TRUE",
    "LastUpdated", "StableWeeks",
]

SCORE_COLUMN_MAP = {
    "Metacritic": "metascore",
    "st.Metacritic": "st_metacritic",
    "Reviews": "review_count",
    "Letterboxd": "letterboxd_rating",
    "st.Letterboxd": "st_letterboxd",
    "IMDB": "imdb_rating",
    "st.IMDB": "st_imdb",
    "TRUE": "composite",
}

_STABILITY_THRESHOLD = 0.05


def load_workbook_from_path(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return wb, ws


def get_header_map(ws) -> dict:
    """Return {header_name: col_index} from the first row."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def ensure_headers(ws, header_map: dict) -> dict:
    """Add any missing output columns to the worksheet."""
    max_col = ws.max_column
    for header in EXPECTED_HEADERS:
        if header not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=header)
            header_map[header] = max_col
    return header_map


def migrate_stability_columns(ws, header_map: dict) -> dict:
    """
    Move LastUpdated and StableWeeks into the columns immediately after TRUE
    so they fall inside Table1 and sort with the rest of the data.

    If they are already adjacent to TRUE (or TRUE is not present), this is a
    no-op.  When a migration is needed the old cells are cleared after copying.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import TableColumn

    true_col = header_map.get("TRUE")
    lu_col = header_map.get("LastUpdated")
    sw_col = header_map.get("StableWeeks")

    if true_col is None or lu_col is None or sw_col is None:
        return header_map

    target_lu = true_col + 1
    target_sw = true_col + 2

    if lu_col == target_lu and sw_col == target_sw:
        return header_map

    max_row = ws.max_row

    table = ws.tables.get(_TABLE_NAME)
    if table is not None:
        new_ref = f"A1:{get_column_letter(target_sw)}{max_row}"
        table.ref = new_ref
        existing = {c.name: c for c in table.tableColumns}
        new_cols = []
        for col_name in _TABLE_CORE_COLS:
            if col_name in existing:
                new_cols.append(existing[col_name])
            else:
                col_idx = header_map.get(col_name) or (
                    target_lu if col_name == "LastUpdated" else target_sw
                )
                new_cols.append(TableColumn(id=col_idx, name=col_name))
        table.tableColumns = new_cols

    if lu_col != target_lu:
        ws.cell(row=1, column=target_lu, value="LastUpdated")
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=target_lu, value=ws.cell(row=r, column=lu_col).value)
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=lu_col).value = None
        header_map["LastUpdated"] = target_lu

    if sw_col != target_sw:
        ws.cell(row=1, column=target_sw, value="StableWeeks")
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=target_sw, value=ws.cell(row=r, column=sw_col).value)
        for r in range(1, max_row + 1):
            ws.cell(row=r, column=sw_col).value = None
        header_map["StableWeeks"] = target_sw

    logger.info(
        "Migrated LastUpdated/StableWeeks to cols %d/%d (inside Table1)",
        target_lu, target_sw,
    )
    return header_map


def extend_table_to_stability_cols(ws) -> None:
    """
    Ensure Table1's ref covers LastUpdated and StableWeeks.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import TableColumn

    table = ws.tables.get(_TABLE_NAME)
    if table is None:
        return

    header_map = get_header_map(ws)
    sw_col = header_map.get("StableWeeks")
    if sw_col is None:
        return

    max_row = ws.max_row
    new_ref = f"A1:{get_column_letter(sw_col)}{max_row}"

    if table.ref == new_ref:
        return

    table.ref = new_ref

    existing_names = {c.name for c in table.tableColumns}
    for col_name in _TABLE_CORE_COLS:
        if col_name not in existing_names:
            col_idx = header_map.get(col_name)
            if col_idx is not None:
                table.tableColumns.append(TableColumn(id=col_idx, name=col_name))

    logger.debug("Updated %s ref to %s", _TABLE_NAME, new_ref)


def read_existing_scores(ws, ws_row: int, header_map: dict) -> RawScores:
    """
    Read the current score values from a workbook row into a RawScores object.
    """
    def _int_cell(col_name):
        col = header_map.get(col_name)
        if col is None:
            return None
        val = ws.cell(row=ws_row, column=col).value
        try:
            return int(val) if val is not None else None
        except (ValueError, TypeError):
            return None

    def _float_cell(col_name):
        col = header_map.get(col_name)
        if col is None:
            return None
        val = ws.cell(row=ws_row, column=col).value
        try:
            return float(val) if val is not None else None
        except (ValueError, TypeError):
            return None

    metascore = _int_cell("Metacritic")
    imdb_rating = _float_cell("IMDB")
    review_count_raw = _int_cell("Reviews")
    review_count = review_count_raw if review_count_raw is not None else 0
    letterboxd_rating = _float_cell("Letterboxd")

    return RawScores(
        title="",
        metascore=metascore,
        imdb_rating=imdb_rating,
        review_count=review_count,
        letterboxd_rating=letterboxd_rating,
    )


def _has_missing_scores(ws, ws_row: int, header_map: dict) -> bool:
    """
    Return True if any core score column is blank AND the row has never been
    successfully processed (no LastUpdated date).
    """
    lu_col = header_map.get("LastUpdated")
    if lu_col and ws.cell(row=ws_row, column=lu_col).value is not None:
        return False

    core_cols = ["Metacritic", "Letterboxd", "IMDB", "TRUE"]
    for col_name in core_cols:
        col_idx = header_map.get(col_name)
        if col_idx is None:
            return True
        if ws.cell(row=ws_row, column=col_idx).value is None:
            return True
    return False


def _read_stability(ws, ws_row: int, header_map: dict) -> tuple:
    """
    Read (last_updated: date | None, stable_weeks: int) from the workbook row.
    """
    last_updated = None
    lu_col = header_map.get("LastUpdated")
    if lu_col:
        raw = ws.cell(row=ws_row, column=lu_col).value
        if raw:
            try:
                if isinstance(raw, (datetime, date)):
                    last_updated = raw if isinstance(raw, date) else raw.date()
                else:
                    last_updated = date.fromisoformat(str(raw)[:10])
            except (ValueError, TypeError):
                pass

    stable_weeks = 0
    sw_col = header_map.get("StableWeeks")
    if sw_col:
        raw = ws.cell(row=ws_row, column=sw_col).value
        try:
            stable_weeks = int(raw) if raw is not None else 0
        except (ValueError, TypeError):
            stable_weeks = 0

    return last_updated, stable_weeks


def should_update(ws, ws_row: int, header_map: dict, today: date) -> bool:
    """
    Return True if this movie should be fetched in a smart-update run.

    Rules:
      1. Any missing core score → always update.
      2. Never been updated (no LastUpdated) → always update.
      3. StableWeeks == 0 → always update.
      4. Otherwise: update only if days since last update >= StableWeeks * 7.
    """
    if _has_missing_scores(ws, ws_row, header_map):
        return True

    last_updated, stable_weeks = _read_stability(ws, ws_row, header_map)

    if last_updated is None or stable_weeks == 0:
        return True

    days_since = (today - last_updated).days
    return days_since >= stable_weeks * 7


def update_stability(
    ws,
    ws_row: int,
    header_map: dict,
    new_composite: Optional[float],
    today: date,
    manual_unchanged: bool = False,
) -> None:
    """
    Update LastUpdated and StableWeeks for a row after a successful fetch.

    StableWeeks increments by 1 if the new composite is within ±0.05 of the
    previous value; resets to 0 if it changed more than that.
    """
    lu_col = header_map.get("LastUpdated")
    sw_col = header_map.get("StableWeeks")

    true_col = header_map.get("TRUE")
    prev_composite = None
    if true_col:
        raw = ws.cell(row=ws_row, column=true_col).value
        try:
            prev_composite = float(raw) if raw is not None else None
        except (ValueError, TypeError):
            pass

    _, prev_stable_weeks = _read_stability(ws, ws_row, header_map)

    if manual_unchanged:
        new_stable_weeks = prev_stable_weeks + 1
    elif new_composite is None or prev_composite is None:
        new_stable_weeks = 0
    elif abs(new_composite - prev_composite) <= _STABILITY_THRESHOLD:
        new_stable_weeks = prev_stable_weeks + 1
    else:
        new_stable_weeks = 0

    if lu_col:
        ws.cell(row=ws_row, column=lu_col, value=today.isoformat())
    if sw_col:
        ws.cell(row=ws_row, column=sw_col, value=new_stable_weeks)
