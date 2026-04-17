#!/usr/bin/env python3
"""
update_scores.py
================
Reads Movies.xlsx, fetches the latest Metacritic / Letterboxd / IMDB scores
for every movie, recalculates the composite score, and writes the results to
Movies_updated.xlsx (the original file is never overwritten).

Usage
-----
    python update_scores.py                        # update all movies
    python update_scores.py --limit 10             # only 10 random movies (testing)
    python update_scores.py --movie "Boogie Nights" # single movie
    python update_scores.py --input my_list.xlsx   # custom input file
    python update_scores.py --delay 1.5            # seconds between requests
    python update_scores.py --api-key YOUR_KEY     # OMDb API key
    python update_scores.py --smart-update         # skip recently-stable movies
    python update_scores.py --manual               # prompt for missing values

Output columns added / updated
-------------------------------
    Metacritic      - Metascore (0-100)
    st.Metacritic   - normalised 0-1
    Reviews         - number of critic reviews
    Letterboxd      - average rating (0-5)
    st.Letterboxd   - normalised 0-1
    IMDB            - IMDB rating (0-10)
    st.IMDB         - normalised 0-1
    TRUE            - composite score (weighted average of the three normalised scores)
    LastUpdated     - ISO date of last successful fetch (YYYY-MM-DD)
    StableWeeks     - consecutive weeks the composite score has been within ±0.05
"""

import argparse
import logging
import os
import random
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from tqdm import tqdm

from scraper.letterboxd_scraper import get_letterboxd_data
from scraper.metacritic_scraper import get_metacritic_data
from scraper.omdb_client import get_omdb_data

# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class RawScores:
    title: str
    metascore: Optional[int]           # 0-100; 50 when OMDb returns N/A
    imdb_rating: Optional[float]       # 0.0-10.0; None when N/A
    review_count: int                  # >= 0; 0 when Metacritic not found
    letterboxd_rating: Optional[float] # 0.0-5.0; None when not found


@dataclass
class NormalisedScores:
    title: str
    metascore: Optional[int]
    st_metacritic: Optional[float]   # 0.0-1.0 or None
    review_count: int
    letterboxd_rating: Optional[float]
    st_letterboxd: Optional[float]   # 0.0-1.0 or None
    imdb_rating: Optional[float]
    st_imdb: Optional[float]         # 0.0-1.0 or None
    composite: Optional[float]       # 0.0-1.0 or None, rounded to 2dp


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Score normalisation helpers (Pass 2)
# ---------------------------------------------------------------------------

def normalise_column(values: list[Optional[float]]) -> list[Optional[float]]:
    """
    Apply min-max normalisation to a column of values.

    - Returns 0.0 for all entries when max == min (flat column).
    - Returns None for entries where the input is None.
    - min and max are computed over non-None values only.
    """
    non_none = [v for v in values if v is not None]
    if not non_none:
        return [None if v is None else 0.0 for v in values]

    col_min = min(non_none)
    col_max = max(non_none)

    if col_max == col_min:
        return [None if v is None else 0.0 for v in values]

    return [
        None if v is None else (v - col_min) / (col_max - col_min)
        for v in values
    ]


def normalise_all(raw_scores: list[RawScores]) -> list[NormalisedScores]:
    """
    Pass 2: apply normalise_column to Metacritic, Letterboxd, and IMDB columns.

    Accepts a list of RawScores dataclasses and returns a list of
    NormalisedScores dataclasses (composite is set to None here; it is
    computed in Pass 3).
    """
    meta_col = [float(r.metascore) if r.metascore is not None else None for r in raw_scores]
    lb_col = [r.letterboxd_rating for r in raw_scores]
    imdb_col = [r.imdb_rating for r in raw_scores]

    norm_meta = normalise_column(meta_col)
    norm_lb = normalise_column(lb_col)
    norm_imdb = normalise_column(imdb_col)

    result = []
    for i, raw in enumerate(raw_scores):
        result.append(NormalisedScores(
            title=raw.title,
            metascore=raw.metascore,
            st_metacritic=norm_meta[i],
            review_count=raw.review_count,
            letterboxd_rating=raw.letterboxd_rating,
            st_letterboxd=norm_lb[i],
            imdb_rating=raw.imdb_rating,
            st_imdb=norm_imdb[i],
            composite=None,
        ))
    return result


# ---------------------------------------------------------------------------
# Composite score helpers (Pass 3)
# ---------------------------------------------------------------------------

def compute_global_anchors(normalised: list[NormalisedScores]) -> tuple[Optional[float], Optional[float]]:
    """
    Compute (Global_Max_St, Global_Min_St) across all non-None values in
    st_metacritic, st_letterboxd, and st_imdb columns combined.

    Returns (None, None) when there are no non-None values.
    """
    all_values = []
    for row in normalised:
        for field in (row.st_metacritic, row.st_letterboxd, row.st_imdb):
            if field is not None:
                all_values.append(field)

    if not all_values:
        return (None, None)

    return (max(all_values), min(all_values))


def compute_composite(
    st_meta: Optional[float],
    reviews: int,
    st_lb: Optional[float],
    st_imdb: Optional[float],
    global_max: Optional[float],
    global_min: Optional[float],
) -> Optional[float]:
    """
    Compute composite score with dynamic denominator.

    Formula (full):
        ((st_meta x reviews) + st_lb + global_max + global_min + st_imdb)
        / (reviews + 4)

    Dynamic adjustments:
        - reviews == 0 or None  -> drop st_meta x reviews term; base denom = 4
        - st_lb is None         -> drop st_lb; denom -= 1
        - st_imdb is None       -> drop st_imdb; denom -= 1
        - global_max or global_min is None -> drop both anchor terms; denom -= 2
        - effective denom == 0  -> return None
    """
    numerator = 0.0
    denominator = 0

    if st_meta is not None and reviews:
        numerator += st_meta * reviews
        denominator += reviews

    if st_lb is not None:
        numerator += st_lb
        denominator += 1

    if global_max is not None and global_min is not None:
        numerator += global_max + global_min
        denominator += 2

    if st_imdb is not None:
        numerator += st_imdb
        denominator += 1

    if denominator == 0:
        return None

    return round(numerator / denominator, 2)


def compute_all_composites(normalised: list[NormalisedScores]) -> list[NormalisedScores]:
    """
    Pass 3: compute composite scores for all movies.

    Calls compute_global_anchors once, then compute_composite per movie.
    Returns a new list of NormalisedScores with the composite field populated.
    """
    global_max, global_min = compute_global_anchors(normalised)

    result = []
    for row in normalised:
        composite = compute_composite(
            st_meta=row.st_metacritic,
            reviews=row.review_count,
            st_lb=row.st_letterboxd,
            st_imdb=row.st_imdb,
            global_max=global_max,
            global_min=global_min,
        )
        result.append(NormalisedScores(
            title=row.title,
            metascore=row.metascore,
            st_metacritic=row.st_metacritic,
            review_count=row.review_count,
            letterboxd_rating=row.letterboxd_rating,
            st_letterboxd=row.st_letterboxd,
            imdb_rating=row.imdb_rating,
            st_imdb=row.st_imdb,
            composite=composite,
        ))
    return result


# ---------------------------------------------------------------------------
# Pass 1: Fetch all raw scores
# ---------------------------------------------------------------------------

def read_existing_scores(ws, ws_row: int, header_map: dict) -> RawScores:
    """
    Read the current score values from a workbook row into a RawScores object.

    Used by apply_manual_entry to compare new manual entries against what is
    already stored.  When the user enters the same values as last time the
    composite won't change, so StableWeeks should increment rather than reset.
    """
    def _int_cell(col_name):
        col = header_map.get(col_name)
        if col is None:
            return None
        val = ws.cell(row=ws_row, column=col).value
        try:
            return int(val) if val is not None else None
        except (ValueError, TypeError):
            return None

    def _float_cell(col_name):
        col = header_map.get(col_name)
        if col is None:
            return None
        val = ws.cell(row=ws_row, column=col).value
        try:
            return float(val) if val is not None else None
        except (ValueError, TypeError):
            return None

    metascore = _int_cell("Metacritic")
    imdb_rating = _float_cell("IMDB")
    review_count_raw = _int_cell("Reviews")
    review_count = review_count_raw if review_count_raw is not None else 0
    letterboxd_rating = _float_cell("Letterboxd")

    # Title is filled in by the caller
    return RawScores(
        title="",
        metascore=metascore,
        imdb_rating=imdb_rating,
        review_count=review_count,
        letterboxd_rating=letterboxd_rating,
    )


def fetch_all(
    movies: list[str],
    api_key: str,
    delay: float = 1.0,
    verbose: bool = False,
) -> tuple[list[RawScores], list[str]]:
    """
    Pass 1: fetch raw scores for all movies.

    For each movie title:
      - Logs the title at INFO level before fetching
      - Calls get_omdb_data(title, api_key) -> metascore, imdb_rating
      - Sleeps delay seconds
      - Calls get_review_count(title) -> review_count
      - Sleeps delay seconds
      - Calls get_letterboxd_data(title) -> letterboxd_rating
      - Sleeps delay seconds
      - Catches per-movie exceptions, logs title + exception, continues

    Returns:
        (raw_scores: list[RawScores], failed: list[str])
        where failed contains titles of movies that raised exceptions.
    """
    raw_scores = []
    failed = []

    for title in tqdm(movies, desc="Fetching scores", unit="movie"):
        logger.info("Fetching: %s", title)
        try:
            omdb = get_omdb_data(title, api_key)
            time.sleep(delay)

            mc = get_metacritic_data(title)
            time.sleep(delay)

            lb = get_letterboxd_data(title)
            time.sleep(delay)

            # Prefer the Metascore scraped directly from Metacritic when
            # available (covers the 1–3 review case where OMDb may not have
            # it).  Fall back to the OMDb value otherwise.
            scraped_metascore = mc.get("metascore")
            omdb_metascore = omdb.get("metascore") if omdb.get("imdb_id") else None
            metascore = scraped_metascore if scraped_metascore is not None else omdb_metascore

            raw_scores.append(RawScores(
                title=title,
                metascore=metascore,
                imdb_rating=omdb.get("imdb_rating"),
                review_count=mc.get("review_count", 0),
                letterboxd_rating=lb.get("rating"),
            ))
        except Exception as exc:
            logger.error("Failed to fetch scores for '%s': %s", title, exc)
            failed.append(title)
            continue

    return raw_scores, failed


# ---------------------------------------------------------------------------
# Manual entry helpers
# ---------------------------------------------------------------------------

def _prompt_value(prompt: str, parser, label: str):
    """
    Prompt the user for a value, parse it with *parser*, and return the result.

    Returns None if the user presses Enter without typing anything (skip).
    Loops until a valid value is entered or the user skips.
    """
    while True:
        raw = input(prompt).strip()
        if raw == "":
            return None
        try:
            return parser(raw)
        except (ValueError, TypeError):
            print(f"  Invalid {label}. Press Enter to skip, or try again.")


def _prompt_int_in_range(prompt: str, lo: int, hi: int, label: str) -> Optional[int]:
    """Prompt for an integer in [lo, hi], returning None on empty input."""
    def parse(s):
        v = int(s)
        if not (lo <= v <= hi):
            raise ValueError(f"{v} not in [{lo}, {hi}]")
        return v
    return _prompt_value(prompt, parse, label)


def _prompt_float_in_range(prompt: str, lo: float, hi: float, label: str) -> Optional[float]:
    """Prompt for a float in [lo, hi], returning None on empty input."""
    def parse(s):
        v = float(s)
        if not (lo <= v <= hi):
            raise ValueError(f"{v} not in [{lo}, {hi}]")
        return v
    return _prompt_value(prompt, parse, label)


def prompt_missing_scores(raw: RawScores) -> RawScores:
    """
    Interactively prompt the user to fill in any None/zero fields on *raw*.

    Called only when --manual is active and a movie has missing data.
    Returns a new RawScores with user-supplied values merged in.
    """
    print(f"\n  ── Manual entry for: {raw.title} ──")
    print("  (Press Enter to skip a field and leave it unchanged)\n")

    metascore = raw.metascore
    imdb_rating = raw.imdb_rating
    review_count = raw.review_count
    letterboxd_rating = raw.letterboxd_rating

    if metascore is None:
        metascore = _prompt_int_in_range(
            "  Metascore (0-100): ", 0, 100, "Metascore"
        )

    if imdb_rating is None:
        imdb_rating = _prompt_float_in_range(
            "  IMDB rating (0.0-10.0): ", 0.0, 10.0, "IMDB rating"
        )

    if review_count == 0:
        rc = _prompt_int_in_range(
            "  Critic review count (0+): ", 0, 100_000, "review count"
        )
        if rc is not None:
            review_count = rc

    if letterboxd_rating is None:
        letterboxd_rating = _prompt_float_in_range(
            "  Letterboxd rating (0.0-5.0): ", 0.0, 5.0, "Letterboxd rating"
        )

    return RawScores(
        title=raw.title,
        metascore=metascore,
        imdb_rating=imdb_rating,
        review_count=review_count,
        letterboxd_rating=letterboxd_rating,
    )


def prompt_failed_movie(title: str) -> Optional[RawScores]:
    """
    Interactively prompt the user to enter all scores for a movie that
    failed entirely during fetch.

    Returns a RawScores if the user provides at least one value, or None
    if the user skips all fields.
    """
    print(f"\n  ── Manual entry for failed movie: {title} ──")
    print("  (Press Enter to skip a field)\n")

    metascore = _prompt_int_in_range(
        "  Metascore (0-100): ", 0, 100, "Metascore"
    )
    imdb_rating = _prompt_float_in_range(
        "  IMDB rating (0.0-10.0): ", 0.0, 10.0, "IMDB rating"
    )
    rc = _prompt_int_in_range(
        "  Critic review count (0+): ", 0, 100_000, "review count"
    )
    review_count = rc if rc is not None else 0
    letterboxd_rating = _prompt_float_in_range(
        "  Letterboxd rating (0.0-5.0): ", 0.0, 5.0, "Letterboxd rating"
    )

    # If the user skipped everything, return None
    if all(v is None for v in (metascore, imdb_rating, letterboxd_rating)) and review_count == 0:
        logger.info("Skipped manual entry for '%s'", title)
        return None

    return RawScores(
        title=title,
        metascore=metascore,
        imdb_rating=imdb_rating,
        review_count=review_count,
        letterboxd_rating=letterboxd_rating,
    )


def _manual_matches_existing(new: RawScores, prev: RawScores) -> bool:
    """
    Return True if every score field in *new* that was manually entered
    matches the corresponding field already stored in *prev*.

    A field is considered "manually entered" (and therefore worth comparing)
    when it is not None / non-zero in *new*.  If *new* has a value and it
    equals *prev*, the entry is unchanged.  We use exact equality for ints
    and a small epsilon for floats to tolerate rounding.
    """
    _EPS = 1e-9

    def _eq(a, b) -> bool:
        if a is None or b is None:
            return a == b
        if isinstance(a, float) or isinstance(b, float):
            return abs(float(a) - float(b)) < _EPS
        return a == b

    # Only compare fields that were actually filled in (non-None / non-zero)
    if new.metascore is not None and not _eq(new.metascore, prev.metascore):
        return False
    if new.imdb_rating is not None and not _eq(new.imdb_rating, prev.imdb_rating):
        return False
    if new.review_count and not _eq(new.review_count, prev.review_count):
        return False
    if new.letterboxd_rating is not None and not _eq(new.letterboxd_rating, prev.letterboxd_rating):
        return False
    return True


def apply_manual_entry(
    raw_scores: list[RawScores],
    failed: list[str],
    manual: bool,
    existing: Optional[dict[str, RawScores]] = None,
) -> tuple[list[RawScores], list[str], set[str]]:
    """
    After Pass 1, optionally prompt the user for missing values.

    - For movies with partial data (None fields), prompts to fill gaps.
    - For movies that failed entirely, prompts to enter all scores.

    When *existing* is provided (a dict mapping title -> RawScores read from
    the workbook before this run), each manually-entered value is compared
    against the stored value.  If every field the user entered matches what
    was already in the workbook, the title is added to the returned
    *manual_unchanged* set.  The caller uses this set to tell update_stability
    to treat the movie as stable (increment StableWeeks) rather than resetting
    it, because the scores haven't actually changed.

    Returns:
        (raw_scores, failed, manual_unchanged)
        where manual_unchanged is a set of titles whose manual entries were
        identical to the existing workbook values.
    """
    manual_unchanged: set = set()

    if not manual:
        return raw_scores, failed, manual_unchanged

    existing = existing or {}

    # Fill gaps in partially-fetched movies
    updated_raw = []
    for raw in raw_scores:
        has_missing = (
            raw.metascore is None
            or raw.imdb_rating is None
            or raw.review_count == 0
            or raw.letterboxd_rating is None
        )
        if has_missing:
            filled = prompt_missing_scores(raw)
            # Check whether the user's entries match the existing workbook values
            prev = existing.get(raw.title)
            if prev is not None and _manual_matches_existing(filled, prev):
                manual_unchanged.add(raw.title)
            raw = filled
        updated_raw.append(raw)

    # Prompt for fully-failed movies
    still_failed = []
    for title in failed:
        result = prompt_failed_movie(title)
        if result is not None:
            prev = existing.get(title)
            if prev is not None and _manual_matches_existing(result, prev):
                manual_unchanged.add(title)
            updated_raw.append(result)
        else:
            still_failed.append(title)

    return updated_raw, still_failed, manual_unchanged


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
EXPECTED_HEADERS = [
    "Movies", "Metacritic", "st.Metacritic", "Reviews",
    "Letterboxd", "st.Letterboxd", "IMDB", "st.IMDB", "TRUE",
    "LastUpdated", "StableWeeks",
]

# Column mapping: workbook column name -> NormalisedScores field name
SCORE_COLUMN_MAP = {
    "Metacritic": "metascore",
    "st.Metacritic": "st_metacritic",
    "Reviews": "review_count",
    "Letterboxd": "letterboxd_rating",
    "st.Letterboxd": "st_letterboxd",
    "IMDB": "imdb_rating",
    "st.IMDB": "st_imdb",
    "TRUE": "composite",
}

# Composite change threshold below which a week is counted as "stable"
_STABILITY_THRESHOLD = 0.05


def load_workbook_from_path(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return wb, ws


def get_header_map(ws) -> dict:
    """Return {header_name: col_index} from the first row."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def ensure_headers(ws, header_map: dict) -> dict:
    """Add any missing output columns to the worksheet."""
    max_col = ws.max_column
    for header in EXPECTED_HEADERS:
        if header not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=header)
            header_map[header] = max_col
    return header_map


# ---------------------------------------------------------------------------
# Smart-update: stability-based scheduling
# ---------------------------------------------------------------------------

def _has_missing_scores(ws, ws_row: int, header_map: dict) -> bool:
    """Return True if any core score column is blank for this row."""
    core_cols = ["Metacritic", "Letterboxd", "IMDB", "TRUE"]
    for col_name in core_cols:
        col_idx = header_map.get(col_name)
        if col_idx is None:
            return True
        if ws.cell(row=ws_row, column=col_idx).value is None:
            return True
    return False


def _read_stability(ws, ws_row: int, header_map: dict) -> tuple:
    """
    Read (last_updated: date | None, stable_weeks: int) from the workbook row.
    """
    last_updated = None
    lu_col = header_map.get("LastUpdated")
    if lu_col:
        raw = ws.cell(row=ws_row, column=lu_col).value
        if raw:
            try:
                if isinstance(raw, (datetime, date)):
                    last_updated = raw if isinstance(raw, date) else raw.date()
                else:
                    last_updated = date.fromisoformat(str(raw)[:10])
            except (ValueError, TypeError):
                pass

    stable_weeks = 0
    sw_col = header_map.get("StableWeeks")
    if sw_col:
        raw = ws.cell(row=ws_row, column=sw_col).value
        try:
            stable_weeks = int(raw) if raw is not None else 0
        except (ValueError, TypeError):
            stable_weeks = 0

    return last_updated, stable_weeks


def should_update(ws, ws_row: int, header_map: dict, today: date) -> bool:
    """
    Return True if this movie should be fetched in a smart-update run.

    Rules:
      1. Any missing core score → always update.
      2. Never been updated (no LastUpdated) → always update.
      3. StableWeeks == 0 → always update.
      4. Otherwise: update only if days since last update >= StableWeeks * 7.
    """
    if _has_missing_scores(ws, ws_row, header_map):
        return True

    last_updated, stable_weeks = _read_stability(ws, ws_row, header_map)

    if last_updated is None or stable_weeks == 0:
        return True

    days_since = (today - last_updated).days
    return days_since >= stable_weeks * 7


def update_stability(
    ws,
    ws_row: int,
    header_map: dict,
    new_composite: Optional[float],
    today: date,
    manual_unchanged: bool = False,
) -> None:
    """
    Update LastUpdated and StableWeeks for a row after a successful fetch.

    StableWeeks increments by 1 if the new composite is within ±0.05 of the
    previous value; resets to 0 if it changed more than that.

    When *manual_unchanged* is True the scores were entered manually and are
    identical to the values already in the workbook, so the composite cannot
    have changed.  StableWeeks is incremented unconditionally in that case,
    matching the behaviour of a scrape that returned the same values as before.
    """
    lu_col = header_map.get("LastUpdated")
    sw_col = header_map.get("StableWeeks")

    # Read previous composite to decide stability
    true_col = header_map.get("TRUE")
    prev_composite = None
    if true_col:
        raw = ws.cell(row=ws_row, column=true_col).value
        try:
            prev_composite = float(raw) if raw is not None else None
        except (ValueError, TypeError):
            pass

    _, prev_stable_weeks = _read_stability(ws, ws_row, header_map)

    # Determine new StableWeeks
    if manual_unchanged:
        # Scores are identical to last run — always stable
        new_stable_weeks = prev_stable_weeks + 1
    elif new_composite is None or prev_composite is None:
        new_stable_weeks = 0
    elif abs(new_composite - prev_composite) <= _STABILITY_THRESHOLD:
        new_stable_weeks = prev_stable_weeks + 1
    else:
        new_stable_weeks = 0

    if lu_col:
        ws.cell(row=ws_row, column=lu_col, value=today.isoformat())
    if sw_col:
        ws.cell(row=ws_row, column=sw_col, value=new_stable_weeks)


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

def update_workbook(
    input_path: Path,
    output_path: Path,
    api_key: str,
    limit: Optional[int] = None,
    target_movie: Optional[str] = None,
    delay: float = 1.0,
    verbose: bool = False,
    smart_update: bool = False,
    manual: bool = False,
):
    """
    Three-pass pipeline:
      Pass 1 - fetch_all: fetch raw scores for all movies
      Pass 2 - normalise_all: column-wide min-max normalisation
      Pass 3 - compute_all_composites: compute composite scores
    Then write results to output workbook.

    When smart_update=True, movies are skipped if their scores have been
    stable long enough (StableWeeks * 7 days since last update).
    Movies with missing scores are always updated.

    When manual=True, the user is prompted to enter any values that could
    not be fetched automatically.  Existing cell values are never overwritten
    by None — if a scraper returns nothing and the user skips manual entry,
    the previous value in the workbook is preserved.
    """
    wb, ws = load_workbook_from_path(input_path)
    header_map = get_header_map(ws)

    title_col = header_map.get("Movies")
    if title_col is None:
        logger.error("Could not find 'Movies' column in %s", input_path)
        sys.exit(1)

    # Ensure all output columns (including LastUpdated, StableWeeks) exist
    header_map = ensure_headers(ws, header_map)

    today = date.today()

    # Collect (worksheet_row_number, title) pairs, skipping blank rows
    movie_rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        title_cell = row[title_col - 1]
        title = title_cell.value
        if title is None or str(title).strip() == "":
            continue
        movie_rows.append((title_cell.row, str(title).strip()))

    # Apply --movie filter
    if target_movie:
        movie_rows = [(r, t) for r, t in movie_rows if t == target_movie]
        if not movie_rows:
            logger.error("Movie '%s' not found in spreadsheet.", target_movie)
            sys.exit(1)

    # Apply --limit filter: pick N movies at random
    if limit:
        random.shuffle(movie_rows)
        movie_rows = movie_rows[:limit]

    # Apply smart-update filter: skip movies that don't need updating yet
    if smart_update:
        skipped = []
        filtered_rows = []
        for ws_row, title in movie_rows:
            if should_update(ws, ws_row, header_map, today):
                filtered_rows.append((ws_row, title))
            else:
                skipped.append(title)
        if skipped:
            logger.info(
                "Smart-update: skipping %d stable movie(s): %s",
                len(skipped),
                ", ".join(skipped),
            )
        movie_rows = filtered_rows

    if not movie_rows:
        logger.info("Nothing to update.")
        wb.save(output_path)
        return

    # Build ordered list of titles for fetch_all
    movies = [t for _, t in movie_rows]

    # Read existing workbook scores before fetching — used by apply_manual_entry
    # to detect when a manual entry is identical to the previously stored value.
    existing_scores: dict = {}
    if manual:
        for ws_row, title in movie_rows:
            prev = read_existing_scores(ws, ws_row, header_map)
            prev.title = title
            existing_scores[title] = prev

    # Pass 1: fetch raw scores for all movies
    raw_scores, failed = fetch_all(movies, api_key=api_key, delay=delay, verbose=verbose)

    # Manual entry: prompt for any values that couldn't be fetched automatically.
    # This happens after all network fetches complete, before normalisation.
    # Returns a third value: set of titles whose manual entries matched existing.
    raw_scores, failed, manual_unchanged = apply_manual_entry(
        raw_scores, failed, manual=manual, existing=existing_scores
    )

    # Pass 2: normalise column-wide (must happen after all fetches complete)
    normalised = normalise_all(raw_scores)

    # Pass 3: compute composite scores
    final_scores = compute_all_composites(normalised)

    # Build lookup: title -> NormalisedScores
    scores_by_title = {ns.title: ns for ns in final_scores}

    # Write results back to workbook rows
    for ws_row, title in movie_rows:
        ns = scores_by_title.get(title)
        if ns is None:
            # Movie failed during fetch — leave row unchanged
            continue
        for col_name, field_name in SCORE_COLUMN_MAP.items():
            value = getattr(ns, field_name)
            if value is None:
                # Leave cell unchanged when value is None
                continue
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=ws_row, column=col_idx, value=value)

        # Update stability tracking columns
        is_unchanged = title in manual_unchanged
        update_stability(ws, ws_row, header_map, ns.composite, today, manual_unchanged=is_unchanged)

    wb.save(output_path)
    logger.info("Saved updated workbook to %s", output_path)

    # Log failed movies summary
    if failed:
        logger.warning("Failed to fetch scores for %d movie(s):", len(failed))
        for t in failed:
            logger.warning("  - %s", t)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Fetch latest Metacritic / Letterboxd / IMDB scores and update Movies.xlsx"
    )
    parser.add_argument(
        "--input", default="Movies.xlsx",
        help="Path to the input Excel file (default: Movies.xlsx)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Path for the output Excel file (default: <input_stem>_updated.xlsx)"
    )
    parser.add_argument(
        "--limit", type=int, default=None,
        help="Pick N movies at random to process (useful for testing)"
    )
    parser.add_argument(
        "--movie", default=None,
        help="Only update a single movie by exact title"
    )
    parser.add_argument(
        "--delay", type=float, default=1.0,
        help="Seconds to wait between requests to each source (default: 1.0)"
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Enable debug logging"
    )
    parser.add_argument(
        "--api-key", default=None,
        dest="api_key",
        help="OMDb API key (overrides OMDB_API_KEY environment variable)"
    )
    parser.add_argument(
        "--smart-update", action="store_true", dest="smart_update",
        help=(
            "Skip movies whose scores have been stable recently. "
            "A movie stable for N weeks is skipped for N weeks. "
            "Movies with missing scores are always updated."
        )
    )
    parser.add_argument(
        "--manual", action="store_true", dest="manual",
        help=(
            "Prompt for manual entry when scores cannot be fetched automatically. "
            "Existing values in the workbook are preserved when a field is skipped."
        )
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Resolve API key: --api-key first, then OMDB_API_KEY env var
    api_key = args.api_key or os.environ.get("OMDB_API_KEY")
    if not api_key:
        logger.error(
            "No OMDb API key provided. Set OMDB_API_KEY environment variable "
            "or pass --api-key."
        )
        sys.exit(1)

    input_path = Path(args.input)
    if not input_path.exists():
        logger.error("Input file not found: %s", input_path)
        sys.exit(1)

    output_path = Path(args.output) if args.output else (
        input_path.parent / f"{input_path.stem}_updated{input_path.suffix}"
    )

    logger.info("Input:  %s", input_path)
    logger.info("Output: %s", output_path)

    update_workbook(
        input_path=input_path,
        output_path=output_path,
        api_key=api_key,
        limit=args.limit,
        target_movie=args.movie,
        delay=args.delay,
        verbose=args.verbose,
        smart_update=args.smart_update,
        manual=args.manual,
    )


if __name__ == "__main__":
    main()
